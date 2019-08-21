import os
from datetime import datetime

import openpyxl
from django.http import HttpResponse
# Create your views here.
from django.shortcuts import render
from openpyxl import Workbook

from college.forms import CloneForm
from .models import AssignSubjectTeacher, Teacher, Expert, Year, Batch, Programme


def generate_xlsx(request, assignsubjectteacherlist="none"):
    if assignsubjectteacherlist == "none":
        assignsubjectteacherlist = AssignSubjectTeacher.objects.all()

    if not request.user.is_superuser:
        return HttpResponse('Sorry! You need to be authorized to access this link.'
                            + "<a href='../'>" + 'Login here.' +
                            "</a> <br> <br> <br> If you think this is a mistake contact WebAdmin.")

    file_path = os.path.join(os.path.dirname(os.path.realpath(__name__)), 'python_spreadsheet.xlsx')
    output_path = os.path.join(os.path.dirname(os.path.realpath(__name__)), 'temp_python_spreadsheet.xlsx')
    wb = Workbook()
    book = openpyxl.load_workbook(file_path)

    sheet = book.get_sheet_by_name('Sheet1')

    row = 4
    col = 0

    sheet['M1'] = datetime.today().strftime('%Y-%m-%d')

    for subjectteachers in assignsubjectteacherlist:
        sheet[colnum_string(col + 1) + str(row)] = str(subjectteachers.subject.subject_code)
        sheet[colnum_string(col + 2) + str(row)] = str(subjectteachers.subject.name)
        # teacher id
        sheet[colnum_string(col + 3) + str(row)] = str(subjectteachers.getyearpart())
        sheet[colnum_string(col + 4) + str(row)] = str(subjectteachers.subject_teacher.teacher_id)
        #
        # # teacher experience required
        sheet[colnum_string(col + 5) + str(row)] = str(subjectteachers.subject_teacher.full_name())
        #
        sheet[colnum_string(col + 6) + str(row)] = str(subjectteachers.subject_teacher.get_teacher_experience_years())
        #
        sheet[colnum_string(col + 7) + str(row)] = str(subjectteachers.subject_teacher_teaching_experience_years)
        #
        sheet[colnum_string(col + 8) + str(row)] = str(subjectteachers.subject_teacher.home_phone)
        #
        sheet[colnum_string(col + 9) + str(row)] = str(subjectteachers.subject_teacher.mobile_phone)
        #
        sheet[colnum_string(col + 10) + str(row)] = str(subjectteachers.subject_teacher.email)
        #
        sheet[colnum_string(col + 11) + str(row)] = str(subjectteachers.subject_teacher.affiliated_institute.code)
        #
        sheet[colnum_string(col + 12) + str(row)] = str(subjectteachers.subject_teacher.upper_degree)
        #
        sheet[colnum_string(col + 13) + str(row)] = str(subjectteachers.subject_teacher.aff_type)

        # TODO add other methods to dump in xlsx here

        row += 1

    book.save(output_path)

    response = HttpResponse(open(output_path, 'rb').read())
    response['Content-Type'] = 'mimetype/submimetype'
    response['Content-Disposition'] = 'attachment; filename=SubjectTeachers.xlsx'
    return response


def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


def exportform(request):
    if not request.user.is_superuser:
        return HttpResponse('Sorry! You need to be authorized to access this link.'
                            + "<a href='../'>" + 'Login here.' +
                            "</a> <br> <br> <br> If you think this is a mistake contact WebAdmin.")

    from college.forms import ExportForm
    if request.method == "POST":

        form = ExportForm(request.POST)

        if form.is_valid():
            # post = form.save(commit=False)
            # post.author = request.user
            # post.date_of_creation = datetime.now()
            # post.save()

            queryList = AssignSubjectTeacher.objects.all()

            if form.cleaned_data['year'] is not None:
                queryList = queryList.filter(year=form.cleaned_data['year'])

            if form.cleaned_data['part'] is not "":
                if form.cleaned_data['part'] == 'Odd':

                    queryList = queryList.filter(semester__in=['First', 'Third'])
                elif form.cleaned_data['part'] == 'Even':

                    queryList = queryList.filter(semester__in=['Second', 'Fourth'])

            if form.cleaned_data['batch'] is not None:
                queryList = queryList.filter(batch=form.cleaned_data['batch'])
            if form.cleaned_data['programme'] is not None:
                queryList = queryList.filter(
                    batch__programme=form.cleaned_data['programme'])
            if form.cleaned_data['semester'] is not "":
                queryList = queryList.filter(semester=form.cleaned_data['semester'])

            return generate_xlsx(request, queryList)

    else:

        form = ExportForm()

    return render(request, 'admin/exportform.html', {'form': form})


def exportteachers(request):
    if not request.user.is_superuser:
        return HttpResponse('Sorry! You need to be authorized to access this link.'
                            + "<a href='../'>" + 'Login here.' +
                            "</a> <br> <br> <br> If you think this is a mistake contact WebAdmin.")
    output_path = os.path.join(os.path.dirname(os.path.realpath(__name__)), 'temp_python_spreadsheet.xlsx')
    wb = Workbook()
    book = openpyxl.Workbook()

    sheet = book.get_sheet_by_name("Sheet")

    row = 3
    col = -4

    sheet['A2'] = "Full Name"
    sheet['B2'] = "Experience Years"
    sheet['C2'] = "Home Phone"
    sheet['D2'] = "Mobile"
    sheet['E2'] = "Email"
    sheet['F2'] = "Affiliated Institute"
    sheet['G2'] = "Upper Degree"
    sheet['H2'] = "Affilation Type"

    sheet['B1'] = "Teachers List"
    sheet['E1'] = "Generated: "
    sheet['F1'] = datetime.today().strftime('%Y-%m-%d')

    for eachteacher in Teacher.objects.all():
        sheet[colnum_string(col + 5) + str(row)] = str(eachteacher.full_name())
        #
        sheet[colnum_string(col + 6) + str(row)] = str(eachteacher.get_teacher_experience_years())
        #
        sheet[colnum_string(col + 7) + str(row)] = str(eachteacher.home_phone)
        #
        sheet[colnum_string(col + 8) + str(row)] = str(eachteacher.mobile_phone)
        #
        sheet[colnum_string(col + 9) + str(row)] = str(eachteacher.email)
        #
        sheet[colnum_string(col + 10) + str(row)] = str(eachteacher.affiliated_institute.code)
        #
        sheet[colnum_string(col + 11) + str(row)] = str(eachteacher.upper_degree)
        #
        sheet[colnum_string(col + 12) + str(row)] = str(eachteacher.aff_type)

        # TODO add other methods to dump in xlsx here

        row += 1

    book.save(output_path)

    response = HttpResponse(open(output_path, 'rb').read())
    response['Content-Type'] = 'mimetype/submimetype'
    response['Content-Disposition'] = 'attachment; filename=TeachersList.xlsx'
    return response


def exportexperts(request):
    if not request.user.is_superuser:
        return HttpResponse('Sorry! You need to be authorized to access this link.'
                            + "<a href='../'>" + 'Login here.' +
                            "</a> <br> <br> <br> If you think this is a mistake contact WebAdmin.")
    output_path = os.path.join(os.path.dirname(os.path.realpath(__name__)), 'temp_python_spreadsheet.xlsx')
    wb = Workbook()
    book = openpyxl.Workbook()

    sheet = book.get_sheet_by_name("Sheet")

    row = 3
    col = -4

    sheet['A2'] = "Full Name"
    sheet['B2'] = "Known Topics"
    sheet['C2'] = "Home Phone"
    sheet['D2'] = "Mobile"
    sheet['E2'] = "Email"
    sheet['F2'] = "Organization "
    sheet['G2'] = "Upper Degree"

    sheet['B1'] = "Experts List"
    sheet['E1'] = "Generated: "
    sheet['F1'] = datetime.today().strftime('%Y-%m-%d')

    for eachteacher in Expert.objects.all():
        sheet[colnum_string(col + 5) + str(row)] = str(eachteacher.full_name())

        sheet[colnum_string(col + 6) + str(row)] = str(eachteacher.get_known_topics())
        #
        sheet[colnum_string(col + 7) + str(row)] = str(eachteacher.home_phone)
        #
        sheet[colnum_string(col + 8) + str(row)] = str(eachteacher.mobile_phone)
        #
        sheet[colnum_string(col + 9) + str(row)] = str(eachteacher.email)
        #
        sheet[colnum_string(col + 10) + str(row)] = str(eachteacher.organization.institute_name)
        #
        sheet[colnum_string(col + 11) + str(row)] = str(eachteacher.upper_degree)

        # TODO add other methods to dump in xlsx here

        row += 1

    book.save(output_path)

    response = HttpResponse(open(output_path, 'rb').read())
    response['Content-Type'] = 'mimetype/submimetype'
    response['Content-Disposition'] = 'attachment; filename=ExpertsList.xlsx'
    return response


def cloneyear(request):
    log = None

    if not request.user.is_superuser:
        return HttpResponse('Sorry! You need to be authorized to access this link.'
                            + "<a href='../'>" + 'Login here.' +
                            "</a> <br> <br> <br> If you think this is a mistake contact WebAdmin.")

    if request.method == "POST":

        form = CloneForm(request.POST)

        if form.is_valid():
            # post = form.save(commit=False)
            # post.author = request.user
            # post.date_of_creation = datetime.now()
            # post.save()

            queryList = AssignSubjectTeacher.objects.all()

            if form.cleaned_data['Academic_year_from'] is not None:
                if form.cleaned_data['Academic_year_to'] is not None:
                    fromyear = str(form.cleaned_data['Academic_year_from'])
                    toyear = str(form.cleaned_data['Academic_year_to'])

                    queryList = queryList.filter(year__name=fromyear)

                    if form.cleaned_data['Semester_type'] is not "":
                        if form.cleaned_data['Semester_type'] == 'Odd':
                            queryList = queryList.filter(semester__in=['First', 'Third'])
                        elif form.cleaned_data['Semester_type'] == 'Even':
                            queryList = queryList.filter(semester__in=['Second', 'Fourth'])

                    if form.cleaned_data['programme'] is not None:
                        queryList = queryList.filter(
                            batch__programme=form.cleaned_data['programme'])

                    log = ""
                    log += "Found " + str(queryList.count()) + ' data<br>'
                    log += 'Cloning data from ' + fromyear + " to " + toyear + '. . . : <br><br>'
                    for object in queryList:
                        object.year = Year.objects.get(name=toyear)

                        teachingyears = object.subject_teacher_teaching_experience_years

                        if teachingyears == "":
                            teachingyears = 0

                        newteachingyears = int(teachingyears) + int(toyear) - int(fromyear)
                        if newteachingyears <= 0:
                            newteachingyears = 0

                        newbatchyear = str(int(object.batch.year.name) + int(toyear) - int(fromyear))

                        Year.objects.get_or_create(name=newbatchyear)

                        Batch.objects.get_or_create(year=Year.objects.get(name=newbatchyear),
                                                    programme=Programme.objects.get(
                                                        pk=object.batch.programme.pk))

                        object.batch = Batch.objects.get(year=Year.objects.get(name=newbatchyear),
                                                         programme=Programme.objects.get(
                                                             pk=object.batch.programme.pk))

                        object.subject_teacher_teaching_experience_years = str(newteachingyears)
                        object.pk = None
                        object.save()
                        log += '- ' + object.subject_teacher.first_name + ', ' + object.subject.name + ', ' + object.batch.year.name + ', ' + object.batch.programme.name + ', ' + object.semester + "<br>"

                    log += ' <br>Completed <br>'

            return render(request, 'admin/yearclone.html', {'form': form, 'log': log})

    else:

        form = CloneForm()

    return render(request, 'admin/yearclone.html', {'form': form, 'log': log})
